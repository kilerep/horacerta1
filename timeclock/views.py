import csv
from datetime import datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

from accounts.models import User
from accounts.mei_context import resolve_mei_context
from companies.models import CompanyAttendancePolicy, CompanyAuthorizedLocation
from .models import ActivityReportRequest, Contract, Punch
from .services import build_daily_summary, evaluate_punch_confidence, filter_punches_by_period, format_hhmm
from .state import contract_operational_q, employee_lifecycle_summary

QR_PRESENCE_SESSION_KEY = "hc_qr_presence_claims"


def _only_employee(user):
    return getattr(user, "role", None) == User.Role.FUNCIONARIO


def _contract_employee_label(contract):
    employee = getattr(contract, "employee", None)
    if not employee:
        return "MEI indisponivel"

    full_name = (getattr(employee, "full_name", "") or "").strip()
    if full_name:
        return full_name

    user = getattr(employee, "user", None)
    if user:
        email = (getattr(user, "email", "") or "").strip()
        if email:
            return email
        username = (getattr(user, "username", "") or "").strip()
        if username:
            return username

    return "MEI indisponivel"


def _active_contracts_for_employee_user(user):
    return (
        Contract.objects.filter(
            employee__user=user,
            employee__is_active=True,
            employee__user__is_active=True,
            employee__isnull=False,
            employee__user__isnull=False,
        )
        .filter(contract_operational_q())
        .select_related("company", "employee", "employee__user")
        .order_by("-start_date", "-created_at")
    )


def _resolve_contract_for_secure_action(request):
    """
    Resolve contrato para acoes sensiveis (exportacoes etc.).
    Se o usuario informar ?contract invalido, nao faz fallback silencioso.
    """
    mei_context = resolve_mei_context(request, operational_only=True)
    if mei_context.requested_contract_id and mei_context.invalid_requested_contract:
        return None, "Vinculo informado invalido ou indisponivel."
    return mei_context.selected_contract, None


def _resolve_selected_contract(contracts_qs, contract_id):
    if not contracts_qs.exists():
        return None
    if contract_id:
        selected = contracts_qs.filter(id=contract_id).first()
        if selected:
            return selected
    return contracts_qs.first()


def _parse_geo_decimal(raw_value):
    value = (raw_value or "").strip()
    if not value:
        return None
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError):
        return None


def _load_qr_claims(session):
    claims = session.get(QR_PRESENCE_SESSION_KEY)
    if isinstance(claims, dict):
        return claims
    return {}


def _save_qr_claims(session, claims):
    session[QR_PRESENCE_SESSION_KEY] = claims
    session.modified = True


def _resolve_qr_requirement(contract, now_local):
    policy = CompanyAttendancePolicy.objects.filter(company=contract.company).first()
    if not policy:
        return {"required": False, "reason": "not_qr_mode"}
    qr_enabled = bool(policy.require_qr) or policy.validation_mode == CompanyAttendancePolicy.ValidationMode.PRESENTIAL_QR
    if not qr_enabled:
        return {"required": False, "reason": "not_qr_mode"}
    if policy.qr_requirement == CompanyAttendancePolicy.QrRequirement.NONE:
        return {"required": False, "reason": "qr_not_required"}

    today = now_local.date()
    day_start = timezone.make_aware(datetime.combine(today, time.min))
    day_end = timezone.make_aware(datetime.combine(today, time.max))
    punches_count_today = Punch.objects.filter(contract=contract, timestamp__range=(day_start, day_end)).count()

    if policy.qr_requirement == CompanyAttendancePolicy.QrRequirement.FIRST_PUNCH:
        return {"required": punches_count_today == 0, "reason": "first_punch", "punches_count_today": punches_count_today}

    # V1: primeira marcacao + fechamento esperado no fim do expediente.
    is_first_punch = punches_count_today == 0
    is_evening_closure = punches_count_today >= 1 and (punches_count_today % 2 == 1) and now_local.hour >= 17
    return {
        "required": is_first_punch or is_evening_closure,
        "reason": "first_and_last",
        "punches_count_today": punches_count_today,
    }


def _policy_audit_snapshot(contract):
    policy = CompanyAttendancePolicy.objects.filter(company=contract.company).first()
    if not policy:
        return {
            "policy_mode": CompanyAttendancePolicy.ValidationMode.FREE,
            "require_location": False,
            "require_qr": False,
            "qr_requirement": CompanyAttendancePolicy.QrRequirement.NONE,
            "default_allowed_radius_m": 120,
            "default_location_id": "",
        }
    return {
        "policy_mode": policy.validation_mode,
        "require_location": bool(policy.require_location),
        "require_qr": bool(policy.require_qr),
        "qr_requirement": policy.qr_requirement,
        "default_allowed_radius_m": int(policy.default_allowed_radius_m or 120),
        "default_location_id": str(policy.default_location_id or ""),
    }


def _consume_valid_qr_claim(request, contract):
    claims = _load_qr_claims(request.session)
    claim = claims.get(str(contract.id))
    if not isinstance(claim, dict):
        return None

    location_id = (claim.get("location_id") or "").strip()
    claimed_at_raw = (claim.get("claimed_at") or "").strip()
    if not location_id or not claimed_at_raw:
        return None

    try:
        claimed_at = datetime.fromisoformat(claimed_at_raw)
    except ValueError:
        return None
    if timezone.is_naive(claimed_at):
        claimed_at = timezone.make_aware(claimed_at, timezone.get_current_timezone())

    if timezone.now() - claimed_at > timedelta(minutes=12):
        return None

    location = CompanyAuthorizedLocation.objects.filter(
        id=location_id,
        company=contract.company,
        is_active=True,
    ).first()
    if not location:
        return None

    claims.pop(str(contract.id), None)
    _save_qr_claims(request.session, claims)
    return location


@login_required
def employee_dashboard(request):
    if not _only_employee(request.user):
        return redirect("dashboard")

    mei_context = resolve_mei_context(request, operational_only=True)
    contracts = mei_context.contracts
    selected_contract = mei_context.selected_contract

    pending_report_requests_qs = ActivityReportRequest.objects.filter(
        employee__user=request.user,
        is_answered=False,
    ).select_related("company", "requested_by", "employee", "employee__user")
    if selected_contract:
        pending_report_requests_qs = pending_report_requests_qs.filter(
            employee=selected_contract.employee,
            company=selected_contract.company,
        )
    pending_report_requests = pending_report_requests_qs[:20]

    if request.method == "POST" and request.POST.get("action") == "respond_activity_request":
        request_id = (request.POST.get("request_id") or "").strip()
        response_text = (request.POST.get("response_text") or "").strip()

        report_request = get_object_or_404(
            ActivityReportRequest,
            id=request_id,
            employee__user=request.user,
            is_answered=False,
        )
        if response_text:
            report_request.response_text = response_text
            report_request.is_answered = True
            report_request.responded_at = timezone.now()
            report_request.save(update_fields=["response_text", "is_answered", "responded_at"])

        redirect_url = f"{request.path}?event=report_sent"
        if selected_contract:
            redirect_url = f"{redirect_url}&contract={selected_contract.id}"
        return redirect(redirect_url)

    if not contracts.exists():
        employee = mei_context.selected_employee
        state_context = employee_lifecycle_summary(employee, []) if employee else None
        employee_company_name = ""
        if employee and getattr(employee, "company", None):
            employee_company_name = employee.company.name
        return render(
            request,
            "accounts/dashboard_funcionario.html",
            {
                "no_contracts": True,
                "contracts": [],
                "state_context": state_context,
                "employee_company_name": employee_company_name,
                "pending_report_requests": pending_report_requests,
                "context_warning": (
                    "O vinculo selecionado anteriormente nao esta mais disponivel. Selecione um vinculo ativo."
                    if (mei_context.invalid_requested_contract or mei_context.invalid_session_contract)
                    else ""
                ),
            },
        )

    if request.method == "POST" and request.POST.get("action") == "punch":
        geo_latitude = _parse_geo_decimal(request.POST.get("geo_latitude"))
        geo_longitude = _parse_geo_decimal(request.POST.get("geo_longitude"))
        geo_accuracy_m = _parse_geo_decimal(request.POST.get("geo_accuracy_m"))
        now_local = timezone.localtime()
        policy_snapshot = _policy_audit_snapshot(selected_contract)
        qr_requirement = _resolve_qr_requirement(selected_contract, now_local)
        qr_required = bool(qr_requirement.get("required"))
        qr_location = _consume_valid_qr_claim(request, selected_contract) if qr_required else None
        qr_status = (
            Punch.QrConfirmationStatus.CONFIRMED
            if qr_location
            else (Punch.QrConfirmationStatus.REQUIRED_MISSING if qr_required else Punch.QrConfirmationStatus.NOT_REQUIRED)
        )
        confidence = evaluate_punch_confidence(
            selected_contract,
            latitude=geo_latitude,
            longitude=geo_longitude,
            accuracy_m=geo_accuracy_m,
        )
        Punch.objects.create(
            contract=selected_contract,
            timestamp=timezone.now(),
            geo_latitude=geo_latitude,
            geo_longitude=geo_longitude,
            geo_accuracy_m=geo_accuracy_m,
            validated_location=confidence.get("validated_location"),
            distance_to_location_m=confidence.get("distance_to_location_m"),
            validation_method=confidence.get("validation_method") or Punch.ValidationMethod.FREE_POLICY,
            confidence_status=confidence.get("confidence_status") or Punch.ConfidenceStatus.FREE,
            qr_confirmation_status=qr_status,
            qr_confirmed_location=qr_location,
            qr_confirmed_at=timezone.now() if qr_location else None,
            audit_payload={
                "source": "WEB_PROFESSIONAL_DASHBOARD",
                "recorded_from": "button_punch",
                "policy": policy_snapshot,
                "qr_required_for_punch": qr_required,
                "qr_requirement_reason": qr_requirement.get("reason") or "",
                "geolocation_collected": bool(geo_latitude is not None and geo_longitude is not None),
                "geolocation_accuracy_m": float(geo_accuracy_m) if geo_accuracy_m is not None else None,
                "request_user_agent": (request.META.get("HTTP_USER_AGENT") or "")[:180],
            },
        )
        if qr_required and not qr_location:
            return redirect(f"{request.path}?event=punch_saved_qr_missing&contract={selected_contract.id}")
        return redirect(f"{request.path}?event=punch_saved&contract={selected_contract.id}")

    date_from_raw = request.GET.get("date_from") or request.GET.get("start")
    date_to_raw = request.GET.get("date_to") or request.GET.get("end")

    qs = Punch.objects.filter(contract=selected_contract)

    today = timezone.localdate()
    start_today = timezone.make_aware(datetime.combine(today, time.min))
    end_today = timezone.make_aware(datetime.combine(today, time.max))
    punches_today = qs.filter(timestamp__range=(start_today, end_today)).order_by("timestamp")

    qs_filtered, date_from, date_to = filter_punches_by_period(qs, date_from_raw, date_to_raw)
    if not date_from and not date_to:
        first_day = today.replace(day=1)
        start_dt = timezone.make_aware(datetime.combine(first_day, time.min))
        end_dt = timezone.make_aware(datetime.combine(today, time.max))
        qs_filtered = qs.filter(timestamp__range=(start_dt, end_dt))

    punches_today_list = list(punches_today)
    total_punches_today = len(punches_today_list)
    today_summary, _today_columns = build_daily_summary(punches_today_list, min_punch_columns=4)
    status_today = today_summary[0]["status"] if today_summary else "INCOMPLETO"
    today_total_partial_hhmm = today_summary[0]["total_hours_hhmm"] if today_summary else "00:00"
    last_punch_today = punches_today_list[-1] if punches_today_list else None
    last_punch_today_label = timezone.localtime(last_punch_today.timestamp).strftime("%H:%M") if last_punch_today else "-"
    today_punch_times = [timezone.localtime(punch.timestamp).strftime("%H:%M") for punch in punches_today_list]
    now_local = timezone.localtime()
    current_hour = now_local.hour
    if 5 <= current_hour <= 11:
        greeting = "Bom dia"
    elif 12 <= current_hour <= 17:
        greeting = "Boa tarde"
    else:
        greeting = "Boa noite"
    day_status_label = "Dia fechado" if total_punches_today % 2 == 0 else "Dia em andamento"
    if total_punches_today == 0:
        journey_status_key = "no_records"
        journey_status_label = "Sem registros hoje"
        journey_status_tone = "neutral"
        journey_next_action = "Registre o primeiro horario do dia para iniciar a jornada."
    elif total_punches_today % 2 == 1 and current_hour >= 20:
        journey_status_key = "incomplete"
        journey_status_label = "Dia incompleto"
        journey_status_tone = "warn"
        journey_next_action = "Dia encerrado com horario pendente. Ajustes operacionais devem ser tratados com o encarregado."
    elif total_punches_today % 2 == 1:
        journey_status_key = "in_progress"
        journey_status_label = "Jornada em andamento"
        journey_status_tone = "progress"
        journey_next_action = "Registre o proximo horario ao concluir a etapa atual da jornada."
    else:
        journey_status_key = "finished"
        journey_status_label = "Dia finalizado"
        journey_status_tone = "ok"
        journey_next_action = "Jornada do dia fechada. Acompanhe o historico e os totais para conferencia."

    history_filtered = list(qs_filtered.order_by("timestamp"))
    history_days, history_punch_columns = build_daily_summary(history_filtered, min_punch_columns=4)
    history_days_desc = sorted(history_days, key=lambda row: row["date"], reverse=True)
    recent_history_days = history_days_desc[:7]
    recent_inconsistencies = [row for row in history_days_desc[:20] if row["is_incomplete"]][:5]

    context = {
        "contracts": contracts,
        "selected_contract": selected_contract,
        "punches_today": punches_today,
        "total_punches_today": total_punches_today,
        "status_today": status_today,
        "greeting": greeting,
        "today_date": now_local.date(),
        "day_status_label": day_status_label,
        "journey_status_key": journey_status_key,
        "journey_status_label": journey_status_label,
        "journey_status_tone": journey_status_tone,
        "journey_next_action": journey_next_action,
        "today_total_partial_hhmm": today_total_partial_hhmm,
        "last_punch_today_label": last_punch_today_label,
        "today_punch_times": today_punch_times,
        "history": qs_filtered.order_by("-timestamp")[:200],
        "history_days": history_days_desc,
        "recent_history_days": recent_history_days,
        "recent_inconsistencies": recent_inconsistencies,
        "history_punch_columns": range(1, history_punch_columns + 1),
        "date_from": date_from_raw or "",
        "date_to": date_to_raw or "",
        "no_contracts": False,
        "state_context": employee_lifecycle_summary(selected_contract.employee, contracts),
        "pending_report_requests": pending_report_requests,
        "context_warning": (
            "O vinculo selecionado anteriormente nao esta mais disponivel. Exibindo o vinculo ativo atual."
            if (mei_context.invalid_requested_contract or mei_context.invalid_session_contract)
            else ""
        ),
    }
    return render(request, "accounts/dashboard_funcionario.html", context)


@login_required
def qr_presence_checkin(request, token):
    if not _only_employee(request.user):
        return redirect("dashboard")

    location = CompanyAuthorizedLocation.objects.filter(qr_token=token, is_active=True).select_related("company").first()
    if not location:
        return redirect(f"{reverse('employee_dashboard')}?event=qr_invalid")

    contract = (
        _active_contracts_for_employee_user(request.user)
        .filter(company=location.company)
        .select_related("company", "employee", "employee__user")
        .first()
    )
    if not contract:
        return redirect(f"{reverse('employee_dashboard')}?event=qr_contract_missing")

    claims = _load_qr_claims(request.session)
    claims[str(contract.id)] = {
        "location_id": str(location.id),
        "claimed_at": timezone.now().isoformat(),
    }
    _save_qr_claims(request.session, claims)
    return redirect(f"{reverse('employee_dashboard')}?event=qr_confirmed&contract={contract.id}")


@login_required
@require_POST
def create_manual_punches(request):
    if not _only_employee(request.user):
        return JsonResponse({"ok": False, "errors": ["Perfil sem permissao para esta operacao."]}, status=403)

    contract_id = (request.POST.get("contract") or "").strip()
    manual_date_raw = (request.POST.get("manual_date") or "").strip()
    raw_times = request.POST.getlist("times")

    errors = []
    if not contract_id:
        errors.append("Selecione um vinculo.")
    if not manual_date_raw:
        errors.append("Informe a data do lancamento.")

    launch_date = None
    if manual_date_raw:
        try:
            launch_date = datetime.strptime(manual_date_raw, "%Y-%m-%d").date()
        except ValueError:
            errors.append("Data invalida.")

    parsed_times = []
    invalid_times = []
    for raw_value in raw_times:
        value = (raw_value or "").strip()
        if not value:
            continue
        try:
            parsed_times.append(datetime.strptime(value, "%H:%M").time())
        except ValueError:
            invalid_times.append(value)

    if invalid_times:
        errors.append("Horario invalido: %s." % ", ".join(invalid_times))
    if not parsed_times:
        errors.append("Informe pelo menos 1 horario.")

    seen_hm = set()
    duplicated_payload_hm = []
    for value in parsed_times:
        hm = (value.hour, value.minute)
        if hm in seen_hm:
            duplicated_payload_hm.append(f"{value.hour:02d}:{value.minute:02d}")
        else:
            seen_hm.add(hm)
    if duplicated_payload_hm:
        unique_dup = sorted(set(duplicated_payload_hm))
        errors.append("Horarios duplicados no lancamento: %s." % ", ".join(unique_dup))

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    contract = _active_contracts_for_employee_user(request.user).filter(id=contract_id).first()
    if not contract:
        return JsonResponse({"ok": False, "errors": ["Vinculo invalido ou inativo."]}, status=400)

    tz = timezone.get_current_timezone()
    day_start = timezone.make_aware(datetime.combine(launch_date, time.min), tz)
    day_end = timezone.make_aware(datetime.combine(launch_date, time.max), tz)
    existing_hm = {
        (local_ts.hour, local_ts.minute)
        for ts in Punch.objects.filter(contract=contract, timestamp__range=(day_start, day_end)).values_list(
            "timestamp", flat=True
        )
        for local_ts in [timezone.localtime(ts, tz)]
    }

    requested_hm = {(item.hour, item.minute) for item in parsed_times}
    duplicated_existing = sorted(f"{hour:02d}:{minute:02d}" for hour, minute in requested_hm if (hour, minute) in existing_hm)
    if duplicated_existing:
        return JsonResponse(
            {
                "ok": False,
                "errors": [
                    "Ja existe horario para este vinculo/data nos horarios: %s." % ", ".join(duplicated_existing)
                ],
            },
            status=400,
        )

    ordered_hm = sorted(requested_hm)
    policy_snapshot = _policy_audit_snapshot(contract)
    confidence = evaluate_punch_confidence(contract, latitude=None, longitude=None, accuracy_m=None)
    with transaction.atomic():
        for hour, minute in ordered_hm:
            manual_timestamp = timezone.make_aware(
                datetime.combine(launch_date, time(hour=hour, minute=minute)),
                tz,
            )
            Punch.objects.create(
                contract=contract,
                timestamp=manual_timestamp,
                is_manual=True,
                validation_method=confidence.get("validation_method") or Punch.ValidationMethod.FREE_POLICY,
                confidence_status=confidence.get("confidence_status") or Punch.ConfidenceStatus.FREE,
                audit_payload={
                    "source": "WEB_PROFESSIONAL_DASHBOARD",
                    "recorded_from": "manual_batch",
                    "policy": policy_snapshot,
                    "qr_required_for_punch": False,
                    "qr_requirement_reason": "manual_entry",
                    "geolocation_collected": False,
                    "request_user_agent": (request.META.get("HTTP_USER_AGENT") or "")[:180],
                },
            )

    return JsonResponse(
        {
            "ok": True,
            "created_count": len(ordered_hm),
            "contract_id": str(contract.id),
        }
    )


@login_required
def export_csv(request):
    if not _only_employee(request.user):
        return redirect("dashboard")

    contract, error_message = _resolve_contract_for_secure_action(request)
    if not contract:
        return HttpResponse(error_message or "Vinculo invalido para exportacao.", status=400)

    base_punches = Punch.objects.filter(contract=contract).order_by("timestamp")
    punches, _start, _end = filter_punches_by_period(
        base_punches,
        request.GET.get("date_from") or request.GET.get("start"),
        request.GET.get("date_to") or request.GET.get("end"),
    )
    daily_rows, max_punches = build_daily_summary(punches, min_punch_columns=4)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="horacerta_{contract.company.name}.csv"'

    writer = csv.writer(response)
    header = ["Empresa", "Data"]
    for idx in range(1, max_punches + 1):
        header.append(f"horário {idx}")
    header.extend(["Total Horas (HH:MM)", "Status"])
    writer.writerow(header)

    for row in sorted(daily_rows, key=lambda x: x["date"], reverse=True):
        writer.writerow(
            [
                contract.company.name,
                row["date"].strftime("%d/%m/%Y"),
                *row["punch_columns"],
                row["total_hours_hhmm"],
                row["status"],
            ]
        )

    return response


@login_required
def export_default(request):
    return export_pdf(request)


@login_required
def export_xlsx(request):
    if not _only_employee(request.user):
        return redirect("dashboard")

    contract, error_message = _resolve_contract_for_secure_action(request)
    if not contract:
        return HttpResponse(error_message or "Vinculo invalido para exportacao.", status=400)

    base_punches = Punch.objects.filter(contract=contract).order_by("timestamp")
    punches, start_date, end_date = filter_punches_by_period(
        base_punches,
        request.GET.get("date_from") or request.GET.get("start"),
        request.GET.get("date_to") or request.GET.get("end"),
    )
    punches = list(punches)

    wb = Workbook()
    ws = wb.active
    ws.title = "horários"

    headers = ["Funcionario", "Empresa", "Data", "Hora"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    employee_name = _contract_employee_label(contract)

    for punch in punches:
        local_ts = timezone.localtime(punch.timestamp)
        ws.append([employee_name, contract.company.name, local_ts.date(), local_ts.time()])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[2].number_format = "DD/MM/YYYY"
        row[3].number_format = "HH:MM"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12

    daily_rows, _max_cols = build_daily_summary(punches, min_punch_columns=4)
    total_seconds = sum(row["total_seconds"] for row in daily_rows)
    total_hours_hhmm = format_hhmm(total_seconds)

    ws.append([])
    ws.append(["Resumo", "", "", ""])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    period_label = (
        f"{start_date.strftime('%d/%m/%Y')} ate {end_date.strftime('%d/%m/%Y')}"
        if start_date and end_date
        else "Periodo completo"
    )
    ws.append(["Periodo", period_label, "", ""])
    ws.append(["Total de horários", len(punches), "", ""])
    ws.append(["Total de horas", total_hours_hhmm, "", ""])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="horacerta_{contract.company.name}_horários.xlsx"'
    return response


@login_required
def export_pdf(request):
    if not _only_employee(request.user):
        return redirect("dashboard")

    contract, error_message = _resolve_contract_for_secure_action(request)
    if not contract:
        return HttpResponse(error_message or "Vinculo invalido para exportacao.", status=400)

    base_punches = Punch.objects.filter(contract=contract).order_by("timestamp")
    punches, start_date, end_date = filter_punches_by_period(
        base_punches,
        request.GET.get("date_from") or request.GET.get("start"),
        request.GET.get("date_to") or request.GET.get("end"),
    )
    punches = list(punches)

    employee_name = _contract_employee_label(contract)
    period_label = (
        f"{start_date.strftime('%d/%m/%Y')} ate {end_date.strftime('%d/%m/%Y')}"
        if start_date and end_date
        else "Periodo completo"
    )

    daily_rows, max_punches = build_daily_summary(punches, min_punch_columns=4)
    total_seconds = sum(row["total_seconds"] for row in daily_rows)
    total_hours_hhmm = format_hhmm(total_seconds)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("HoraCerta - Resumo de horários", styles["Title"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"<b>Funcionario:</b> {employee_name}", styles["Normal"]))
    story.append(Paragraph(f"<b>Empresa:</b> {contract.company.name}", styles["Normal"]))
    story.append(Paragraph(f"<b>Periodo:</b> {period_label}", styles["Normal"]))
    story.append(Paragraph(f"<b>Total de horas:</b> {total_hours_hhmm}", styles["Normal"]))
    story.append(Paragraph(f"<b>Total de horários:</b> {len(punches)}", styles["Normal"]))
    story.append(Spacer(1, 10))

    table_headers = ["Data"] + [f"horário {idx}" for idx in range(1, max_punches + 1)] + ["Total do dia", "Status"]
    table_data = [table_headers]
    for row in daily_rows:
        status_label = "Incompleto" if row["is_incomplete"] else "OK"
        table_data.append(
            [
                row["date"].strftime("%d/%m/%Y"),
                *row["punch_columns"],
                row["total_hours_hhmm"],
                status_label,
            ]
        )

    if len(table_data) == 1:
        table_data.append(["-"] + ["-"] * max_punches + ["00:00", "OK"])

    date_width = 24 * mm
    total_width = 26 * mm
    status_width = 24 * mm
    usable_width = (landscape(A4)[0] - doc.leftMargin - doc.rightMargin) - (
        date_width + total_width + status_width
    )
    each_punch_width = max(16 * mm, usable_width / max(1, max_punches))
    col_widths = [date_width] + [each_punch_width] * max_punches + [total_width, status_width]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#20396b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#8ea6d1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (1, 1), (max_punches + 2, -1), "CENTER"),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="horacerta_{contract.company.name}_resumo.pdf"'
    return response
